from pathlib import Path
import sys
import types

# Ensure the application package is importable
BASE_DIR = Path(__file__).resolve().parents[1] / "2_Aplikacja_Glowna"
sys.path.insert(0, str(BASE_DIR))

from gui import pdf_processor_app
from gui.pdf_processor_app import PdfProcessorApp, INFO_FIELDS


def test_export_to_xlsx_applies_styles(tmp_path, monkeypatch):
    """Ensure Excel export applies header and status styling."""

    # -- prepare dummy tree data --
    class DummyItem:
        def __init__(self, text="", color="#FFFFFF"):
            self._text = text
            self._color = color

        def text(self):
            return self._text

        def background(self):
            return types.SimpleNamespace(
                color=lambda: types.SimpleNamespace(name=lambda: self._color)
            )

    headers = ["Nazwa pliku", "Lp.", "Nowa nazwa", *[label for _, label in INFO_FIELDS]]
    row = [DummyItem("file.pdf"), DummyItem("1"), DummyItem("new.pdf")]
    row.extend(DummyItem("") for _ in headers[3:-1])
    row.append(DummyItem("OK"))

    class DummyTree:
        def rowCount(self):
            return 1

        def columnCount(self):
            return len(headers)

        def item(self, r, c):
            return row[c]

    app = PdfProcessorApp.__new__(PdfProcessorApp)
    app.tree = DummyTree()

    # -- patch QtWidgets dialogs --
    out_path = tmp_path / "out.xlsx"

    class DummyFileDialog:
        @staticmethod
        def getSaveFileName(*args, **kwargs):
            return str(out_path), ""

    class DummyMessageBox:
        @staticmethod
        def information(*args, **kwargs):
            pass

        @staticmethod
        def critical(*args, **kwargs):
            pass

    monkeypatch.setattr(
        pdf_processor_app,
        "QtWidgets",
        types.SimpleNamespace(QFileDialog=DummyFileDialog, QMessageBox=DummyMessageBox),
    )

    # -- stub openpyxl workbook --
    from openpyxl.utils import get_column_letter

    class DummyCell:
        def __init__(self):
            self.font = None
            self.fill = None
            self.border = None
            self.alignment = None

    class DummyWorksheet:
        def __init__(self, max_col):
            self.max_col = max_col
            self.cells = {}
            self.freeze_panes = None
            self.auto_filter = types.SimpleNamespace(ref=None)
            self.dimensions = "A1"
            import collections

            self.column_dimensions = collections.defaultdict(
                lambda: types.SimpleNamespace(width=None)
            )

        def __getitem__(self, idx):
            return [self.cell(idx, c) for c in range(1, self.max_col + 1)]

        def cell(self, row, column):
            key = (row, column)
            if key not in self.cells:
                self.cells[key] = DummyCell()
            return self.cells[key]

    class DummyWorkbook:
        def __init__(self):
            self.active = DummyWorksheet(len(headers) - 1)
            self.saved = False

        def save(self, path):
            self.saved = True

    dummy_wb = DummyWorkbook()

    import openpyxl

    monkeypatch.setattr(openpyxl, "load_workbook", lambda path: dummy_wb, raising=False)

    # -- execute --
    app.export_to_xlsx()

    # -- assertions --
    header_cell = dummy_wb.active.cell(1, 1)
    assert header_cell.fill.start_color == "D9D9D9"
    assert header_cell.border.left.style == "thin"

    status_col = len(headers) - 1
    status_cell = dummy_wb.active.cell(2, status_col)
    assert status_cell.fill.start_color == "C6EFCE"
