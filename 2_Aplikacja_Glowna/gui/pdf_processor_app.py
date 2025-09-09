"""Qt-based PDF processor application.

This module provides a minimal Qt implementation of the original
``PdfProcessorApp`` which previously relied on ``tkinter``.  The new
version demonstrates how the GUI can be migrated to PySide6 while keeping
selected pieces of business logic such as ``handle_file_copy``.

The GUI is intentionally lightweight â€“ it focuses on showcasing Qt
patterns required by the exercises:

* ``PdfProcessorApp`` now derives from :class:`QtWidgets.QMainWindow`.
* ``StringVar``/``BooleanVar`` usages are replaced with plain attributes
  updated via Qt signals (``QLineEdit.textChanged`` and
  ``QCheckBox.stateChanged``).
* ``ttk`` widgets have been swapped for their PySide6 counterparts.
* All dialog windows utilise :class:`QtWidgets.QFileDialog` and
  :class:`QtWidgets.QMessageBox`.
* Background processing is handled by a :class:`ProcessingWorker`
  subclassing :class:`QtCore.QThread` and emitting progress/error signals.

The application only implements a very small subset of the original
functionality â€“ enough to select a folder with PDF files and list them in
``QTreeWidget`` using a worker thread.  The rest of the original Tk
features can be ported incrementally following the same patterns.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
from pathlib import Path
import inspect
import pandas as pd
from datetime import datetime
import getpass

from .qt_safe import QtWidgets, QtCore, QtGui
from . import processing_worker, style
from .processing_worker import ProcessingWorker
from .training_window import TrainingWindow
from .session_manager_ui import SessionManagerUI
from .constants import DOC_TYPE_LABELS, LABEL_TO_CODE
from config import AppSettings, load_settings, save_settings
from app_session_manager import SessionManager

# Load bundled fonts so they are available across platforms
try:  # pragma: no cover - run only when Qt is available
    _fonts_dir = Path(__file__).with_name("fonts")
    for _font_path in _fonts_dir.glob("*.ttf"):
        QtGui.QFontDatabase.addApplicationFont(str(_font_path))
except Exception:  # pragma: no cover - missing Qt or fonts
    pass

try:  # pragma: no cover - optional PDF preview dependencies
    from PySide6.QtPdf import QPdfDocument
    from PySide6.QtPdfWidgets import QPdfView
except Exception:  # pragma: no cover - provide dummies when import fails
    QPdfDocument = None  # type: ignore
    QPdfView = None  # type: ignore

try:  # pragma: no cover - heavy deps may be missing in tests
    from ..ml_helper import DocumentLLMProcessor
except Exception:  # pragma: no cover - provide dummy when import fails
    DocumentLLMProcessor = None  # type: ignore


class _DummySignal:
    """Fallback signal used when PySide6 is not available."""

    def __init__(self) -> None:  # pragma: no cover - simple stub
        self._slots = []

    def connect(self, slot, *args, **kwargs):  # pragma: no cover - simple stub
        self._slots.append(slot)

    def emit(self, *args, **kwargs):  # pragma: no cover - simple stub
        for slot in list(self._slots):
            slot(*args, **kwargs)


logger = logging.getLogger(__name__)


# Kolumny odpowiadajÄ…ce danym zwracanym przez ``extract_info_from_text``.
# Pierwszy element to klucz w sÅ‚owniku z informacjami, drugi â€“ etykieta
# wyÅ›wietlana w tabeli oraz w eksporcie.
INFO_FIELDS = [
    ("data", "Data"),
    ("nadawca_odbiorca", "Nadawca/Odbiorca"),
    ("w_sprawie", "W sprawie"),
    ("numer_dokumentu", "Numer dokumentu"),
    ("sygnatura_sprawy", "Sygnatura sprawy"),
    ("typ_dokumentu", "Typ dokumentu"),
    ("status", "Status"),
]


class ConfigDialog(QtWidgets.QDialog):
    """Simple dialog allowing the user to edit application settings."""

    def __init__(self, settings: AppSettings, session_manager=None, parent=None):
        super().__init__(parent)
        self.session_manager = session_manager
        self.setWindowTitle("Ustawienia")
        layout = QtWidgets.QFormLayout(self)

        self.tesseract_edit = QtWidgets.QLineEdit(settings.tesseract_folder)
        layout.addRow("Tesseract", self.tesseract_edit)

        self.poppler_edit = QtWidgets.QLineEdit(settings.poppler_folder)
        layout.addRow("Poppler", self.poppler_edit)

        self.output_subdir_edit = QtWidgets.QLineEdit(settings.default_output_subdir)
        layout.addRow("DomyÅ›lny podfolder", self.output_subdir_edit)

        self.language_edit = QtWidgets.QLineEdit(settings.ocr_language)
        layout.addRow("JÄ™zyk OCR", self.language_edit)

        self.psm_spin = QtWidgets.QSpinBox()
        self.psm_spin.setRange(0, 13)
        self.psm_spin.setValue(settings.ocr_psm)
        layout.addRow("PSM", self.psm_spin)

        self.oem_spin = QtWidgets.QSpinBox()
        self.oem_spin.setRange(0, 3)
        self.oem_spin.setValue(settings.ocr_oem)
        layout.addRow("OEM", self.oem_spin)

        self.ocr_dpi_spin = QtWidgets.QSpinBox()
        self.ocr_dpi_spin.setRange(72, 1200)
        self.ocr_dpi_spin.setValue(settings.ocr_dpi)
        layout.addRow("DPI OCR", self.ocr_dpi_spin)

        self.blur_spin = QtWidgets.QSpinBox()
        self.blur_spin.setRange(1, 99)
        self.blur_spin.setSingleStep(2)
        self.blur_spin.setValue(settings.blur_kernel_size)
        layout.addRow("Rozmycie", self.blur_spin)

        self.block_spin = QtWidgets.QSpinBox()
        self.block_spin.setRange(3, 99)
        self.block_spin.setSingleStep(2)
        self.block_spin.setValue(settings.adaptive_threshold_block_size)
        layout.addRow("Rozmiar bloku", self.block_spin)

        self.c_spin = QtWidgets.QSpinBox()
        self.c_spin.setRange(-20, 20)
        self.c_spin.setValue(settings.adaptive_threshold_c)
        layout.addRow("Parametr C", self.c_spin)

        reset_btn = QtWidgets.QPushButton("Resetuj licznik")
        reset_btn.clicked.connect(self._on_reset_counters)
        layout.addRow(reset_btn)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _on_reset_counters(self) -> None:
        if self.session_manager:
            self.session_manager.reset_counters()
            try:
                QtWidgets.QMessageBox.information(
                    self, "Reset", "Liczniki zostaÅ‚y zresetowane."
                )
            except Exception:
                pass
            try:
                self.session_manager.app._sync_number_edit()
            except Exception:
                pass


# Wzorzec niedozwolonych znakÃ³w w nazwach plikÃ³w (w tym znaki sterujÄ…ce)
DISALLOWED_CHARS = re.compile(r"[<>:\"/\\|?*\x00-\x1F]")


def handle_file_copy(src_path: str, dest_dir: str, filename: str) -> str | None:
    r"""Copy ``src_path`` to ``dest_dir`` ensuring the filename is safe.

    ``filename`` is sanitised by replacing any character outside the ASCII
    ``[\w.-]`` range with an underscore.  The resulting name is returned when
    the copy succeeds; ``None`` indicates that the file was skipped or an
    error occurred.
    """

    safe_name = Path(filename).name
    safe_name = re.sub(r"[^\w.-]", "_", safe_name, flags=re.ASCII)

    if DISALLOWED_CHARS.search(safe_name):
        logger.warning(
            "Pomijanie pliku '%s' z powodu niedozwolonych znakÃ³w w nazwie '%s'",
            src_path,
            safe_name,
        )
        return None

    destination = Path(dest_dir) / safe_name
    try:
        shutil.copy2(src_path, destination)
        if safe_name != filename:
            logger.info(
                "Nazwa pliku '%s' zostaÅ‚a zmieniona na '%s'", filename, safe_name
            )

        return safe_name
    except Exception as e:  # pragma: no cover - logujemy bÅ‚Ä…d
        logger.error("BÅ‚Ä…d kopiowania pliku %s: %s", src_path, e)
        return None


class PdfProcessorApp(QtWidgets.QMainWindow, SessionManagerUI):
    """Simplified Qt implementation of the PDF processor."""

    def __init__(self, settings: AppSettings | None = None) -> None:
        super().__init__()
        self.settings = settings or load_settings()
        if hasattr(self, "setWindowTitle"):
            self.setWindowTitle(self.settings.gui_title)
        if hasattr(self, "resize"):
            self.resize(self.settings.gui_width, self.settings.gui_height)

        # Former ``StringVar``/``BooleanVar`` replaced with plain attributes
        self.work_mode: str = "KP"
        self.input_dir: str = ""
        self.output_dir: str = ""
        self.case_signature: str = ""
        self.use_llm: bool = False
        self.llm_processor = None
        # track whether a missing-model warning has already been shown
        self._llm_warning_shown: bool = False

        self._worker: ProcessingWorker | None = None
        self._training_window: TrainingWindow | None = None
        self._highlighted_rows: set[int] = set()
        self.pdf_doc = None
        self.pdf_view = None
        self._current_pdf_path: str | None = None
        self.current_user = getpass.getuser()
        self.current_date = datetime.now().strftime("%Y-%m-%d")

        try:  # skip UI construction when Qt stubs are used in tests
            self._build_ui()
        except Exception:  # pragma: no cover - defensive fallback
            pass

        # Apply saved theme on startup
        self._apply_theme(self.settings.theme)

        # Inicjalizuj menedÅ¼era sesji
        try:
            self.session_manager = SessionManager(self)
        except Exception:
            self.session_manager = None
        try:
            self._sync_number_edit()
        except Exception:
            pass

    # ------------------------------------------------------------------
    # UI construction
    def _build_ui(self) -> None:
        central = QtWidgets.QWidget(self)
        self.setCentralWidget(central)

        outer_layout = QtWidgets.QVBoxLayout(central)

        # button to toggle the side panel
        toggle_layout = QtWidgets.QHBoxLayout()
        toggle_layout.addStretch()
        self.panel_toggle_btn = QtWidgets.QToolButton(text="<<")
        self.panel_toggle_btn.clicked.connect(self._toggle_side_panel)
        toggle_layout.addWidget(self.panel_toggle_btn)
        outer_layout.addLayout(toggle_layout)

        self.main_splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        outer_layout.addWidget(self.main_splitter, 1)

        # --- Left pane: table and PDF preview ---
        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)

        self.info_keys = [key for key, _ in INFO_FIELDS]
        self.table_headers = [
            "Nazwa pliku",
            "Lp.",
            "Nowa nazwa",
            *[label for _, label in INFO_FIELDS],
        ]
        self.tree = QtWidgets.QTableWidget(0, len(self.table_headers))
        self.tree.setHorizontalHeaderLabels(self.table_headers)
        self.tree.setEditTriggers(QtWidgets.QAbstractItemView.AllEditTriggers)
        self.tree.installEventFilter(self)
        self.tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        self.tree.itemChanged.connect(self._update_new_name)
        self.tree.itemChanged.connect(self._on_item_changed)
        self.tree.currentCellChanged.connect(lambda *args: self._update_preview())
        self.tree.itemSelectionChanged.connect(self._update_preview)

        view_splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        view_splitter.addWidget(self.tree)

        if QPdfDocument and QPdfView:
            self.pdf_doc = QPdfDocument(self)
            self.pdf_view = QPdfView()
            self.pdf_view.setDocument(self.pdf_doc)
        else:  # pragma: no cover - fallback when PDF support missing
            self.pdf_doc = None
            self.pdf_view = QtWidgets.QLabel("PodglÄ…d niedostÄ™pny")

        view_splitter.addWidget(self.pdf_view)
        view_splitter.setStretchFactor(0, 3)
        view_splitter.setStretchFactor(1, 2)
        left_layout.addWidget(view_splitter)

        self.main_splitter.addWidget(left_widget)

        # --- Right pane: controls ---
        self.side_panel = QtWidgets.QWidget()
        panel_layout = QtWidgets.QVBoxLayout(self.side_panel)

        # Input directory selection
        in_layout = QtWidgets.QHBoxLayout()
        panel_layout.addLayout(in_layout)
        in_layout.addWidget(QtWidgets.QLabel("Katalog wejÅ›ciowy:"))
        self.input_edit = QtWidgets.QLineEdit()
        self.input_edit.textChanged.connect(self._on_input_dir_changed)
        in_layout.addWidget(self.input_edit, 1)
        browse_in = QtWidgets.QPushButton("Wybierzâ€¦")
        browse_in.clicked.connect(self.select_input_dir)
        in_layout.addWidget(browse_in)

        # Output directory selection
        out_layout = QtWidgets.QHBoxLayout()
        panel_layout.addLayout(out_layout)
        out_layout.addWidget(QtWidgets.QLabel("Katalog wynikowy:"))
        self.output_edit = QtWidgets.QLineEdit()
        self.output_edit.textChanged.connect(self._on_output_dir_changed)
        out_layout.addWidget(self.output_edit, 1)
        browse_out = QtWidgets.QPushButton("Wybierzâ€¦")
        browse_out.clicked.connect(self.select_output_dir)
        out_layout.addWidget(browse_out)

        # Case signature
        sig_layout = QtWidgets.QHBoxLayout()
        panel_layout.addLayout(sig_layout)
        sig_layout.addWidget(QtWidgets.QLabel("Sygnatura:"))
        self.case_edit = QtWidgets.QLineEdit()
        self.case_edit.textChanged.connect(self._on_case_changed)
        sig_layout.addWidget(self.case_edit, 1)
        sig_layout.addWidget(QtWidgets.QLabel("Numer:"))
        self.signature_edit = QtWidgets.QLineEdit()
        self.signature_edit.setFixedWidth(60)
        self.signature_edit.textChanged.connect(self._on_signature_changed)
        sig_layout.addWidget(self.signature_edit)

        # Work mode
        mode_layout = QtWidgets.QHBoxLayout()
        panel_layout.addLayout(mode_layout)
        mode_layout.addWidget(QtWidgets.QLabel("Tryb pracy:"))
        self.mode_combo = QtWidgets.QComboBox()
        self.mode_combo.addItems(DOC_TYPE_LABELS.values())
        self.mode_combo.currentTextChanged.connect(self._on_mode_changed)
        mode_layout.addWidget(self.mode_combo, 1)
        try:
            self.mode_combo.setCurrentText(DOC_TYPE_LABELS[self.work_mode])
        except Exception:
            pass

        try:
            self._sync_number_edit()
        except Exception:
            pass

        # LLM checkbox
        self.llm_checkbox = QtWidgets.QCheckBox("UÅ¼yj LLM")
        self.llm_checkbox.stateChanged.connect(
            lambda state: setattr(self, "use_llm", bool(state))
        )
        panel_layout.addWidget(self.llm_checkbox)

        # Start and validation buttons
        self.start_btn = QtWidgets.QPushButton("Skanuj katalog")
        self.start_btn.clicked.connect(self.start_processing)
        panel_layout.addWidget(self.start_btn)

        self.validate_btn = QtWidgets.QPushButton("Waliduj wiersz")
        self.validate_btn.clicked.connect(self._validate_current_row)
        panel_layout.addWidget(self.validate_btn)

        panel_layout.addStretch(1)

        self.main_splitter.addWidget(self.side_panel)
        self.main_splitter.setStretchFactor(0, 3)
        self.main_splitter.setStretchFactor(1, 1)

        self.side_panel.setVisible(self.settings.panel_visible)
        self.panel_toggle_btn.setText("<<" if self.settings.panel_visible else ">>")

        # --- Menu actions ---
        menu = self.menuBar()
        session_menu = menu.addMenu("Sesja")
        load_action = session_menu.addAction("Wczytaj sesjÄ™")
        load_action.triggered.connect(self.load_existing_session)
        save_action = session_menu.addAction("Zapisz sesjÄ™")
        save_action.triggered.connect(self.save_current_session)

        tools_menu = menu.addMenu("NarzÄ™dzia")
        train_action = tools_menu.addAction("Trening modelu")
        train_action.triggered.connect(self.open_training_window)

        export_action = tools_menu.addAction("Eksportuj do XLSX")
        export_action.triggered.connect(self.export_to_xlsx)
        about_action = menu.addAction("O programie")
        about_action.triggered.connect(self.show_about_dialog)

        settings_menu = menu.addMenu("Ustawienia")
        config_action = settings_menu.addAction("Konfiguracjaâ€¦")
        config_action.triggered.connect(self.edit_settings)

        theme_menu = settings_menu.addMenu("Motyw")
        group = QtGui.QActionGroup(self)
        light_action = theme_menu.addAction("Jasny")
        light_action.setCheckable(True)
        dark_action = theme_menu.addAction("Ciemny")
        dark_action.setCheckable(True)
        group.addAction(light_action)
        group.addAction(dark_action)
        if self.settings.theme == "dark":
            dark_action.setChecked(True)
        else:
            light_action.setChecked(True)
        light_action.triggered.connect(lambda: self._apply_theme("light", save=True))
        dark_action.triggered.connect(lambda: self._apply_theme("dark", save=True))

        # --- Toolbar ---
        toolbar = self.addToolBar("Main")
        add_action = QtGui.QAction("Dodaj plik", self)
        add_action.triggered.connect(self.add_file)
        add_action.setShortcut(QtGui.QKeySequence("Ctrl+O"))
        toolbar.addAction(add_action)

        del_action = QtGui.QAction("UsuÅ„ wiersz", self)
        del_action.triggered.connect(self.remove_current_row)
        del_action.setShortcut(QtGui.QKeySequence("Del"))
        toolbar.addAction(del_action)

        copy_action = QtGui.QAction("Kopiuj dane", self)
        copy_action.triggered.connect(self.copy_selected_data)
        copy_action.setShortcut(QtGui.QKeySequence("Ctrl+C"))
        toolbar.addAction(copy_action)

        self.addActions([add_action, del_action, copy_action])

        # --- Shortcuts for menu actions ---
        load_action.setShortcut(QtGui.QKeySequence("Ctrl+Shift+O"))
        save_action.setShortcut(QtGui.QKeySequence("Ctrl+S"))
        export_action.setShortcut(QtGui.QKeySequence("Ctrl+E"))

        # --- Progress bar in the status bar ---
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setRange(0, 1)
        self.progress_bar.setValue(0)
        try:
            self.statusBar().addPermanentWidget(self.progress_bar)
        except Exception:
            pass

    def _toggle_side_panel(self) -> None:
        """Show or hide the side control panel."""
        visible = not self.side_panel.isVisible()
        try:
            self.side_panel.setVisible(visible)
            self.panel_toggle_btn.setText("<<" if visible else ">>")
        except Exception:
            pass
        self.settings.panel_visible = visible
        save_settings(self.settings)

    def _apply_theme(self, theme: str, save: bool = False) -> None:
        """Apply QSS theme and optionally persist the choice."""
        try:
            self.setStyleSheet(style.get_qss(theme))
        except Exception:  # pragma: no cover - run without Qt
            pass
        self.settings.theme = theme
        if save:
            save_settings(self.settings)

    # ------------------------------------------------------------------
    def open_training_window(self) -> None:
        if self._training_window is None:
            self._training_window = TrainingWindow(self)
            try:
                self._training_window.finished.connect(
                    lambda *args: setattr(self, "_training_window", None)
                )
            except Exception:
                pass
            self._training_window.show()
        else:
            self._training_window.raise_()
            self._training_window.activateWindow()

    def edit_settings(self) -> None:
        """Open configuration dialog and persist changes."""
        dlg = ConfigDialog(self.settings, self.session_manager, self)
        if dlg.exec():
            self.settings.tesseract_folder = dlg.tesseract_edit.text()
            self.settings.poppler_folder = dlg.poppler_edit.text()
            self.settings.default_output_subdir = dlg.output_subdir_edit.text()
            self.settings.ocr_language = dlg.language_edit.text()
            self.settings.ocr_psm = dlg.psm_spin.value()
            self.settings.ocr_oem = dlg.oem_spin.value()
            self.settings.ocr_dpi = dlg.ocr_dpi_spin.value()
            self.settings.blur_kernel_size = dlg.blur_spin.value()
            self.settings.adaptive_threshold_block_size = dlg.block_spin.value()
            self.settings.adaptive_threshold_c = dlg.c_spin.value()
            save_settings(self.settings)
            try:
                from processing import ocr

                ocr._configure_pytesseract()
            except Exception:  # pragma: no cover - optional during tests
                pass

    def show_about_dialog(self) -> None:
        """Display basic author and library information."""
        try:
            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle("O programie")
            layout = QtWidgets.QVBoxLayout(dialog)

            author = QtWidgets.QLabel(
                'Autor: <a href="https://github.com/kitajusSus">Krzysztof Bezubik</a>'
                '<p> Program przygotowany w trakcie staÅ¼u w Izbie Gospodarczej Gazownictwa.</p>'
                '<p> Przekazuje go do uÅ¼ytku wewnÄ™trznego by nie byÅ‚o problemÃ³w licencyjnych (wszyskie biblioteki sÄ… open-source tzn darmowe).</p>'
                '<p> W razie pytaÅ„ lub problemÃ³w proszÄ™ o kontakt mailowy: krzysztof.bezub@gmail.com </p>'
                '<p> Pozdrawiam serdeczenie ðŸ™‚ PaniÄ… dyrektor LaskowskÄ…, AnetÄ™, paniÄ… ElizÄ™, Leszka, Staszka, JulitÄ™, MonikÄ™, Jacka(naczelny kofirejver-er), Wojtka, AgatÄ™, MaÄ‡ka (Å›miaÅ‚ siÄ™ najgÅ‚oÅ›niej z moich Å¼artÃ³w ale bÄ™dzie zaprzeczaÅ‚) i oczywiÅ›cie Matiego </p>'
            )
            author.setOpenExternalLinks(True)
            layout.addWidget(author)
            
            libs = QtWidgets.QLabel(
                "<p>Aplikacja wykorzystuje biblioteki:</p>"
                "<ul>"
                '<li><a href="https://github.com/pyside/PySide6">PySide6</a></li>'
                '<li><a href="https://github.com/pandas-dev/pandas">pandas</a></li>'
                '<li><a href="https://github.com/madmaze/pytesseract">pytesseract</a></li>'
                '<a href="https://dowcipy.jeja.pl/9217,rzecz-sie-dzieje-w-gorach.html">.</a>'    
                "</ul>"
            )
            libs.setOpenExternalLinks(True)
            layout.addWidget(libs)

            btns = QtWidgets.QDialogButtonBox(
                QtWidgets.QDialogButtonBox.Ok, parent=dialog
            )
            btns.accepted.connect(dialog.accept)
            layout.addWidget(btns)
            dialog.exec()
        except Exception:
            pass

    # ------------------------------------------------------------------
    # UI callbacks updating attributes
    def _on_input_dir_changed(self, text: str) -> None:
        self.input_dir = text
        if not text:
            return
        target = Path(text) / self.settings.default_output_subdir
        if not self.output_dir or Path(self.output_dir) != target:
            self.output_edit.setText(str(target))

    def _on_output_dir_changed(self, text: str) -> None:
        self.output_dir = text

    def _on_case_changed(self, text: str) -> None:
        self.case_signature = text

    def _on_mode_changed(self, text: str) -> None:
        self.work_mode = LABEL_TO_CODE.get(text, self.work_mode)
        self._sync_number_edit()

    def _on_signature_changed(self, text: str) -> None:
        if not self.session_manager:
            return
        if not text.isdigit():
            return
        year = datetime.now().year
        counters = (
            self.session_manager.sa_counters
            if self.work_mode == "SA"
            else self.session_manager.counters
        )
        counters[f"{self.work_mode}/{year}"] = max(0, int(text) - 1)

    def _sync_number_edit(self) -> None:
        if not self.session_manager:
            return
        year = datetime.now().year
        counters = (
            self.session_manager.sa_counters
            if self.work_mode == "SA"
            else self.session_manager.counters
        )
        num = counters.get(f"{self.work_mode}/{year}", 0) + 1
        try:
            self.signature_edit.setText(str(num))
        except Exception:
            pass

    # ------------------------------------------------------------------
    # File dialogs
    def select_input_dir(self) -> None:
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Wybierz folder z PDF"
        )
        if directory:
            self.input_edit.setText(directory)

    def select_output_dir(self) -> None:
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Wybierz katalog wynikowy"
        )
        if directory:
            self.output_edit.setText(directory)

    # ------------------------------------------------------------------
    # Processing logic
    def start_processing(self) -> None:
        """Start processing in a background ``QThread``."""

        if not self.input_dir:
            try:
                QtWidgets.QMessageBox.warning(
                    self, "BÅ‚Ä…d", "Wybierz katalog wejÅ›ciowy z plikami PDF."
                )
            except Exception:
                pass
            return

        llm = None
        if self.use_llm:
            llm = self.init_llm_processor()
            if llm is None:
                if not self._llm_warning_shown:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "Brak modelu LLM",
                        "Nie moÅ¼na wÅ‚Ä…czyÄ‡ funkcji LLM â€“ model jest niedostÄ™pny.",
                    )
                    self._llm_warning_shown = True
                return

        self.start_btn.setEnabled(False)
        try:
            self.tree.setRowCount(0)
        except Exception:
            try:
                self.tree.clear()
            except Exception:
                pass

        # Przy kaÅ¼dorazowym uruchomieniu przetwarzania
        # resetuj zestaw kolumn do domyÅ›lnego stanu, aby
        # dynamiczne kolumny z poprzednich uruchomieÅ„
        # nie byÅ‚y przenoszone na nowe sesje.
        self.info_keys = [key for key, _ in INFO_FIELDS]
        self.table_headers = [
            "Nazwa pliku",
            "Lp.",
            "Nowa nazwa",
            *[label for _, label in INFO_FIELDS],
        ]
        if hasattr(self.tree, "setColumnCount"):
            self.tree.setColumnCount(len(self.table_headers))
        if hasattr(self.tree, "setHorizontalHeaderLabels"):
            self.tree.setHorizontalHeaderLabels(self.table_headers)

        out_dir = self.output_dir or self.input_dir
        params = inspect.signature(ProcessingWorker).parameters
        counters = None
        if self.session_manager:
            counters = (
                self.session_manager.sa_counters
                if self.work_mode == "SA"
                else self.session_manager.counters
            )
        kwargs = {
            "input_dir": self.input_dir,
            "output_dir": out_dir,
            "work_mode": self.work_mode,
            "case_signature": self.case_signature,
            "llm_processor": llm,
            "settings": self.settings,
            "counters": counters,
        }
        filtered = {k: v for k, v in kwargs.items() if k in params}
        self._worker = ProcessingWorker(**filtered)
        self._worker.progress.connect(self._on_processing_progress)
        self._worker.finished.connect(self._on_processing_finished)
        self._worker.error.connect(self._on_processing_error)
        self._worker.finished.connect(lambda: self.start_btn.setEnabled(True))
        self._worker.start()

    def _on_processing_finished(self, results: list) -> None:
        self.tree.blockSignals(True)
        # Wykryj dodatkowe klucze w wynikach i dynamicznie dodaj brakujÄ…ce kolumny.
        additional_keys: list[str] = []
        for res in results:
            info = res[3] if len(res) > 3 and isinstance(res[3], dict) else {}
            for key in info.keys():
                if key in ("colors", *self.info_keys) or key in additional_keys:
                    continue
                additional_keys.append(key)

        if additional_keys:
            for key in additional_keys:
                self.info_keys.append(key)
                label = next((lbl for k, lbl in INFO_FIELDS if k == key), None)
                if label is None:
                    label = key.replace("_", " ").title()
                self.table_headers.append(label)
            self.tree.setColumnCount(len(self.table_headers))
            self.tree.setHorizontalHeaderLabels(self.table_headers)

        self.tree.setRowCount(len(results))
        for row, result in enumerate(results):
            # ``process_files`` zwraca krotkÄ™ (nazwa, lp, nowa_nazwa, info)
            name, lp, new_name, info = (
                result[0],
                result[1],
                result[2],
                (result[3] if len(result) > 3 else {}),
            )

            name_item = QtWidgets.QTableWidgetItem(name)
            lp_item = QtWidgets.QTableWidgetItem(str(lp))
            new_item = QtWidgets.QTableWidgetItem(new_name)
            self.tree.setItem(row, 0, name_item)
            self.tree.setItem(row, 1, lp_item)
            self.tree.setItem(row, 2, new_item)

            full_path = os.path.join(self.input_dir, name)
            try:
                name_item.setData(QtCore.Qt.UserRole, full_path)
            except Exception:
                pass

            # UzupeÅ‚nij kolumny informacyjne
            for offset, key in enumerate(self.info_keys, start=3):
                value = info.get(key, "") if isinstance(info, dict) else ""
                self.tree.setItem(row, offset, QtWidgets.QTableWidgetItem(value))

            colors = info.get("colors", {}) if isinstance(info, dict) else {}
            if colors:
                self._highlighted_rows.add(row)
                # zaznacz caÅ‚y wiersz na jasnoÅ¼Ã³Å‚to
                for col in range(self.tree.columnCount()):
                    item = self.tree.item(row, col)
                    if item:
                        item.setBackground(QtGui.QColor(255, 255, 200))
                # wyrÃ³Å¼nij konkretne puste pola
            for key, color in colors.items():
                if key in self.info_keys:
                    col_idx = self.info_keys.index(key) + 3
                    cell = self.tree.item(row, col_idx)
                    if cell:
                        cell.setBackground(QtGui.QColor(color))

        self.tree.blockSignals(False)
        if results:
            try:
                self.tree.selectRow(0)
                self._update_preview()
            except Exception:
                pass
        self._sync_number_edit()

    def _on_processing_error(self, message: str) -> None:
        QtWidgets.QMessageBox.critical(self, "BÅ‚Ä…d", message)

    def _on_processing_progress(self, current: int, total: int) -> None:
        """Update progress information in the status bar."""
        try:
            self.progress_bar.setMaximum(total)
            self.progress_bar.setValue(current)
            self.statusBar().showMessage(f"Przetwarzanie {current}/{total}")
        except Exception:
            pass

    def _load_pdf(self, path: str) -> None:
        if not self.pdf_doc:
            return
        try:
            self.pdf_doc.load(path)
            self._current_pdf_path = path
        except Exception:
            pass

    def _update_preview(self, *_) -> None:
        if not hasattr(self.tree, "currentRow"):
            return
        row = self.tree.currentRow()
        if row < 0:
            return
        item = self.tree.item(row, 0)
        if item is None:
            return
        path = item.data(QtCore.Qt.UserRole)
        if path:
            self._load_pdf(path)

    def _update_new_name(self, item) -> None:
        """Rebuild the value in the "Nowa nazwa" column.

        The filename is constructed from ``Lp.``, ``Data``, ``Typ dokumentu``
        and ``Nadawca/Odbiorca`` columns and sanitised using
        :data:`DISALLOWED_CHARS`. Signals are temporarily blocked while
        updating the target cell to prevent recursive ``itemChanged``
        emissions.
        """

        lp_idx = self.table_headers.index("Lp.")
        date_idx = self.table_headers.index("Data")
        type_idx = self.table_headers.index("Typ dokumentu")
        sender_idx = self.table_headers.index("Nadawca/Odbiorca")

        # Only respond to changes in columns affecting the new filename
        if item.column() not in (lp_idx, date_idx, type_idx, sender_idx):
            return

        row = item.row()
        lp_item = self.tree.item(row, lp_idx)
        date_item = self.tree.item(row, date_idx)
        type_item = self.tree.item(row, type_idx)
        sender_item = self.tree.item(row, sender_idx)
        lp = lp_item.text() if lp_item else ""
        data = date_item.text() if date_item else ""
        typ = type_item.text() if type_item else ""
        nadawca = sender_item.text() if sender_item else ""

        new_name = f"{lp}_{data}_{typ}_{nadawca}".strip("_") + ".pdf"
        new_name = DISALLOWED_CHARS.sub("_", new_name)

        target = self.tree.item(row, 2)
        if target is None:
            target = QtWidgets.QTableWidgetItem()
            self.tree.setItem(row, 2, target)

        self.tree.blockSignals(True)
        target.setText(new_name)
        self.tree.blockSignals(False)

    def _validate_current_row(self) -> None:
        """Re-run analysis for the currently selected row."""
        if not hasattr(self.tree, "currentRow"):
            return
        row = self.tree.currentRow()
        if row < 0:
            return
        name_item = self.tree.item(row, 0)
        if name_item is None:
            return
        path = name_item.data(QtCore.Qt.UserRole)
        if not path:
            path = os.path.join(self.input_dir, name_item.text())

        try:
            text = Path(path).read_text("utf-8", errors="ignore")
        except Exception:
            text = ""

        llm = self.init_llm_processor() if self.use_llm else None
        info = processing_worker.extract_info_from_text(
            text, name_item.text(), self.work_mode, self.case_signature, llm
        )
        try:
            counters = {}
            if self.session_manager:
                counters = (
                    self.session_manager.sa_counters
                    if self.work_mode == "SA"
                    else self.session_manager.counters
                )
            new_name = processing_worker.generate_new_filename(
                info, self.work_mode, counters
            )
            self._sync_number_edit()
        except ValueError:
            new_name = f"dokument_do_weryfikacji_{row+1}.pdf"

        self.tree.blockSignals(True)
        self.tree.setItem(row, 2, QtWidgets.QTableWidgetItem(new_name))
        for offset, key in enumerate(self.info_keys, start=3):
            value = info.get(key, "")
            self.tree.setItem(row, offset, QtWidgets.QTableWidgetItem(value))
        self.tree.blockSignals(False)

        # clear previous highlighting
        for col in range(self.tree.columnCount()):
            cell = self.tree.item(row, col)
            if cell:
                cell.setBackground(QtGui.QBrush())
        self._highlighted_rows.discard(row)

        colors = info.get("colors", {})
        if colors:
            self._highlighted_rows.add(row)
            for col in range(self.tree.columnCount()):
                item = self.tree.item(row, col)
                if item:
                    item.setBackground(QtGui.QColor(255, 255, 200))
            for key, color in colors.items():
                if key in self.info_keys:
                    col_idx = self.info_keys.index(key) + 3
                    cell = self.tree.item(row, col_idx)
                    if cell:
                        cell.setBackground(QtGui.QColor(color))

    def add_file(self) -> None:
        """Append a single PDF file to the table."""
        try:
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "Wybierz plik PDF", self.input_dir, "PDF (*.pdf)"
            )
        except Exception:
            path = ""
        if not path:
            return
        row = self.tree.rowCount()
        self.tree.insertRow(row)
        name = os.path.basename(path)
        item = QtWidgets.QTableWidgetItem(name)
        try:
            item.setData(QtCore.Qt.UserRole, path)
        except Exception:
            pass
        self.tree.setItem(row, 0, item)
        self.tree.selectRow(row)
        self._load_pdf(path)

    def remove_current_row(self) -> None:
        """Remove the currently selected row from the table."""
        if not hasattr(self.tree, "currentRow"):
            return
        row = self.tree.currentRow()
        if row >= 0:
            self.tree.removeRow(row)

    def copy_selected_data(self) -> None:
        """Copy data from the current row to the clipboard."""
        if not hasattr(self.tree, "currentRow"):
            return
        row = self.tree.currentRow()
        if row < 0:
            return
        values: list[str] = []
        for col in range(self.tree.columnCount()):
            item = self.tree.item(row, col)
            values.append(item.text() if item else "")
        try:
            QtWidgets.QApplication.clipboard().setText("\t".join(values))
        except Exception:
            pass

    def _on_item_changed(self, item):
        """Handle manual edits: mark cell green, clear row highlight and update status."""
        status_idx = self.table_headers.index("Status")
        # ignore programmatic changes to status column
        if item.column() == status_idx:
            return

        row = item.row()
        if row in self._highlighted_rows:
            # remove previous highlight from the row
            for col in range(self.tree.columnCount()):
                cell = self.tree.item(row, col)
                if cell:
                    cell.setBackground(QtGui.QBrush())
            self._highlighted_rows.discard(row)

        item.setBackground(QtGui.QColor("lightgreen"))

        status_item = self.tree.item(row, status_idx)
        if status_item:
            self.tree.blockSignals(True)
            status_item.setText("OK")
            self.tree.blockSignals(False)

        if row == self.tree.currentRow():
            name_item = self.tree.item(row, 0)
            if name_item:
                path = name_item.data(QtCore.Qt.UserRole)
                if path == self._current_pdf_path:
                    self._load_pdf(path)

    def _on_tree_item_double_clicked(self, item, column):
        """Open the selected PDF when an item is double-clicked."""
        row = item.row() if hasattr(item, "row") else 0
        name_item = self.tree.item(row, 0) if hasattr(self.tree, "item") else item
        try:
            filepath = name_item.data(QtCore.Qt.UserRole)
        except Exception:
            filepath = None
        if not filepath:
            original = name_item.text().split(" -> ")[0]
            filepath = os.path.join(self.input_dir, original)
        try:
            if not processing_worker.open_pdf_file(filepath):
                QtWidgets.QMessageBox.critical(
                    self, "BÅ‚Ä…d", f"Nie moÅ¼na otworzyÄ‡ pliku PDF:\n{filepath}"
                )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "BÅ‚Ä…d", str(exc))

    def export_to_xlsx(self) -> None:
        """Export table contents to an XLSX file."""
        try:
            path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Eksportuj do XLSX", "", "Pliki Excel (*.xlsx)"
            )
        except Exception:
            path = ""
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        headers = ["Lp.", "Nowa nazwa", *[label for _, label in INFO_FIELDS]]
        data: list[list[str]] = []
        colors: list[list[str]] = []
        for row in range(self.tree.rowCount()):
            row_values: list[str] = []
            row_colors: list[str] = []
            for col in range(1, self.tree.columnCount()):
                item = self.tree.item(row, col)
                row_values.append(item.text() if item else "")
                if item:
                    try:
                        qc = item.background().color().name().upper()
                    except Exception:
                        qc = ""
                    hex_color = qc.lstrip("#")
                    if hex_color in {"00000000", "FFFFFF", "FFFFC8"}:
                        hex_color = ""
                else:
                    hex_color = ""
                row_colors.append(hex_color)
            data.append(row_values)
            colors.append(row_colors)

        col_widths = [len(h) for h in headers]
        for row_values in data:
            for idx, value in enumerate(row_values):
                col_widths[idx] = max(col_widths[idx], len(str(value)))

        df = pd.DataFrame(data, columns=headers)
        try:
            df.to_excel(path, index=False)
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = load_workbook(path)
            ws = wb.active
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            bold = Font(bold=True)
            header_fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.font = bold
                cell.fill = header_fill
                cell.border = border
            for idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width + 2
            alt_colors = ["FFFFFF", "F0F0F0"]
            for r_idx, color_row in enumerate(colors, start=2):
                row_fill = alt_colors[(r_idx - 2) % 2]
                for c_idx, hex_color in enumerate(color_row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
                    fill_color = hex_color or row_fill
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid",
                    )

            status_col = headers.index("Status") + 1
            status_fills = {
                "OK": "C6EFCE",
                "BÅÄ„D": "F8CBAD",
                "DO UZUPEÅNIENIA": "FFF3CD",
            }
            for r_idx, row in enumerate(data, start=2):
                status = row[status_col - 1]
                color = status_fills.get(status)
                if color:
                    cell = ws.cell(row=r_idx, column=status_col)
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

            wb.save(path)
            QtWidgets.QMessageBox.information(
                self, "Eksport", f"Zapisano dane do pliku:\n{path}"
            )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "BÅ‚Ä…d", str(exc))

    def eventFilter(self, source, event):
        tree = getattr(self, "tree", None)
        if tree is source and event.type() == QtCore.QEvent.KeyPress:
            key = event.key()
            modifiers = event.modifiers()
            row = tree.currentRow()
            col = tree.currentColumn()
            if key == QtCore.Qt.Key_Tab:
                if modifiers & QtCore.Qt.ShiftModifier:
                    if col > 0:
                        tree.setCurrentCell(row, col - 1)
                    elif row > 0:
                        tree.setCurrentCell(row - 1, tree.columnCount() - 1)
                else:
                    if col + 1 < tree.columnCount():
                        tree.setCurrentCell(row, col + 1)
                    elif row + 1 < tree.rowCount():
                        tree.setCurrentCell(row + 1, 0)
                return True
            if key == QtCore.Qt.Key_Right:
                if col + 1 < tree.columnCount():
                    tree.setCurrentCell(row, col + 1)
                return True
            if key == QtCore.Qt.Key_Left:
                if col > 0:
                    tree.setCurrentCell(row, col - 1)
                return True
            if key == QtCore.Qt.Key_Down:
                if row + 1 < tree.rowCount():
                    tree.setCurrentCell(row + 1, col)
                return True
            if key == QtCore.Qt.Key_Up:
                if row > 0:
                    tree.setCurrentCell(row - 1, col)
                return True
        return super().eventFilter(source, event)

    # ------------------------------------------------------------------
    def closeEvent(self, event):  # pragma: no cover - simple shutdown logic
        if self._worker and self._worker.isRunning():
            try:
                self._worker.stop()
                self._worker.wait()
            except Exception:
                pass
        super().closeEvent(event)

    def init_llm_processor(self):
        """Initialise and cache an LLM processor if available."""
        if self.llm_processor:
            return self.llm_processor

        if DocumentLLMProcessor is None:
            self._llm_warning_shown = True
            QtWidgets.QMessageBox.warning(
                self,
                "Brak modelu LLM",
                "Brak moduÅ‚u przetwarzania LLM. Zainstaluj model w menedÅ¼erze modeli.",
            )
            try:  # pragma: no cover - GUI may be unavailable in tests
                from .model_manager import ModelManager

                ModelManager(self).exec()
            except Exception:
                pass
            return None
        try:
            processor = DocumentLLMProcessor()
            if hasattr(processor, "load_model") and not processor.load_model():
                raise RuntimeError("model not loaded")
            self.llm_processor = processor
            return processor
        except Exception:  # pragma: no cover - safe guard
            self._llm_warning_shown = True
            QtWidgets.QMessageBox.warning(
                self,
                "Brak modelu LLM",
                "Nie udaÅ‚o siÄ™ wczytaÄ‡ modelu. Zainstaluj model w menedÅ¼erze modeli.",
            )
            try:  # pragma: no cover - GUI may be unavailable in tests
                from .model_manager import ModelManager

                ModelManager(self).exec()
            except Exception:
                pass
            return None

    # ------------------------------------------------------------------
    @staticmethod
    def process_files(
        input_dir: str,
        output_dir: str = "",
        work_mode: str = "KP",
        case_signature: str = "",
        llm_processor=None,
        counters=None,
    ) -> list[tuple[str, int, str, dict]]:
        """Proxy to :func:`processing_worker.process_files` returning detailed info."""
        return processing_worker.process_files(
            input_dir,
            output_dir,
            work_mode,
            case_signature,
            llm_processor,
            counters,
        )


__all__ = ["PdfProcessorApp", "ProcessingWorker", "handle_file_copy"]
